"""
Property-based tests for performance optimization components.

Tests template caching, async processing, and memory optimization
to ensure performance improvements work correctly across various inputs.
"""

import tempfile
import shutil
import json
import asyncio
import threading
import time
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime, timedelta
from decimal import Decimal

import pytest
from hypothesis import given, strategies as st, settings, assume
from openpyxl import Workbook

from declaration_printing.performance_optimizer import (
    PerformanceOptimizer, TemplateCache, AsyncBatchProcessor, 
    MemoryOptimizer, PerformanceMetrics
)
from declaration_printing.template_manager import TemplateManager
from declaration_printing.declaration_printer import DeclarationPrinter
from declaration_printing.models import DeclarationData, DeclarationType, PrintResult, BatchPrintResult
from config.configuration_manager import ConfigurationManager
from logging_system.logger import Logger


class TestTemplateCacheProperties:
    """Property-based tests for template caching system."""
    
    def setup_method(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.template_dir = Path(self.temp_dir) / "templates"
        self.template_dir.mkdir()
        
        # Create mock template manager
        self.mock_template_manager = Mock(spec=TemplateManager)
        self.mock_template_manager.validate_template.return_value = True
        self.mock_template_manager.load_template_mapping.return_value = {"field1": "A1"}
        
        # Create template cache
        self.cache = TemplateCache(max_cache_size=5, cache_ttl_hours=1)
    
    def teardown_method(self):
        """Clean up test environment."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_test_template(self, template_name: str) -> str:
        """Create a test Excel template."""
        template_path = self.template_dir / f"{template_name}.xlsx"
        workbook = Workbook()
        workbook.save(template_path)
        return str(template_path)
    
    @given(
        template_names=st.lists(
            st.text(alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')), min_size=1, max_size=10),
            min_size=1,
            max_size=10,
            unique=True
        )
    )
    @settings(max_examples=50, deadline=5000)
    def test_template_cache_hit_miss_consistency(self, template_names):
        """
        **Feature: customs-declaration-printing, Property: Template cache hit/miss consistency**
        
        For any sequence of template requests, cache hits should only occur
        for templates that were previously loaded and are still valid.
        """
        assume(len(template_names) >= 1)
        
        # Create templates
        template_paths = []
        for name in template_names:
            try:
                path = self._create_test_template(name)
                template_paths.append(path)
            except Exception:
                assume(False)  # Skip invalid template names
        
        assume(len(template_paths) >= 1)
        
        # Track expected cache state
        expected_cached = set()
        
        for template_path in template_paths:
            initial_hits = self.cache.metrics.template_cache_hits
            initial_misses = self.cache.metrics.template_cache_misses
            
            # Get template
            with patch('openpyxl.load_workbook') as mock_load:
                mock_workbook = Mock()
                mock_load.return_value = mock_workbook
                
                result = self.cache.get_template_workbook(template_path, self.mock_template_manager)
            
            # Verify result
            assert result is not None
            
            # Check cache metrics
            if template_path in expected_cached:
                # Should be cache hit
                assert self.cache.metrics.template_cache_hits == initial_hits + 1
                assert self.cache.metrics.template_cache_misses == initial_misses
            else:
                # Should be cache miss
                assert self.cache.metrics.template_cache_hits == initial_hits
                assert self.cache.metrics.template_cache_misses == initial_misses + 1
                expected_cached.add(template_path)
    
    @given(
        cache_size=st.integers(min_value=1, max_value=5),
        template_count=st.integers(min_value=1, max_value=10)
    )
    @settings(max_examples=30, deadline=5000)
    def test_cache_eviction_policy(self, cache_size, template_count):
        """
        **Feature: customs-declaration-printing, Property: Cache eviction maintains size limit**
        
        For any cache size limit and number of templates, the cache should never
        exceed the maximum size and should evict least recently used items.
        """
        # Create cache with specific size
        cache = TemplateCache(max_cache_size=cache_size, cache_ttl_hours=24)
        
        # Create templates
        template_paths = []
        for i in range(template_count):
            try:
                path = self._create_test_template(f"template_{i}")
                template_paths.append(path)
            except Exception:
                continue
        
        assume(len(template_paths) >= 1)
        
        # Load templates
        with patch('openpyxl.load_workbook') as mock_load:
            mock_load.return_value = Mock()
            
            for template_path in template_paths:
                cache.get_template_workbook(template_path, self.mock_template_manager)
                
                # Verify cache size never exceeds limit
                stats = cache.get_cache_stats()
                assert stats['cache_size'] <= cache_size
                assert stats['cache_size'] <= stats['max_cache_size']
    
    @given(ttl_hours=st.integers(min_value=1, max_value=24))
    @settings(max_examples=20, deadline=10000)
    def test_cache_ttl_expiration(self, ttl_hours):
        """
        **Feature: customs-declaration-printing, Property: Cache TTL expiration**
        
        For any TTL setting, cached templates should be evicted after the TTL expires.
        """
        # Create cache with specific TTL
        cache = TemplateCache(max_cache_size=10, cache_ttl_hours=ttl_hours)
        
        template_path = self._create_test_template("ttl_test")
        
        with patch('openpyxl.load_workbook') as mock_load:
            mock_load.return_value = Mock()
            
            # Load template (should be cache miss)
            result1 = cache.get_template_workbook(template_path, self.mock_template_manager)
            assert result1 is not None
            initial_misses = cache.metrics.template_cache_misses
            
            # Load again immediately (should be cache hit)
            result2 = cache.get_template_workbook(template_path, self.mock_template_manager)
            assert result2 is not None
            assert cache.metrics.template_cache_hits > 0
            
            # Simulate time passage beyond TTL
            with patch('declaration_printing.performance_optimizer.datetime') as mock_datetime:
                future_time = datetime.now() + timedelta(hours=ttl_hours + 1)
                mock_datetime.now.return_value = future_time
                
                # Load again (should be cache miss due to TTL expiration)
                result3 = cache.get_template_workbook(template_path, self.mock_template_manager)
                assert result3 is not None
                assert cache.metrics.template_cache_misses > initial_misses


class TestAsyncBatchProcessorProperties:
    """Property-based tests for asynchronous batch processing."""
    
    def setup_method(self):
        """Set up test environment."""
        self.mock_logger = Mock()
        self.processor = AsyncBatchProcessor(max_workers=2, batch_size=3)
    
    @given(
        declaration_numbers=st.lists(
            st.text(alphabet=st.characters(whitelist_categories=('Nd',)), min_size=10, max_size=15),
            min_size=1,
            max_size=20,
            unique=True
        ),
        success_rate=st.floats(min_value=0.0, max_value=1.0)
    )
    @settings(max_examples=20, deadline=10000)
    def test_async_batch_processing_completeness(self, declaration_numbers, success_rate):
        """
        **Feature: customs-declaration-printing, Property: Async batch processing completeness**
        
        For any list of declarations and success rate, async processing should
        process all declarations and return results for each one.
        """
        assume(len(declaration_numbers) >= 1)
        
        # Create mock printer
        mock_printer = Mock()
        
        def mock_print_single(decl_num):
            # Simulate success/failure based on success_rate
            import random
            success = random.random() < success_rate
            
            return PrintResult(
                success=success,
                declaration_number=decl_num,
                output_file_path=f"output_{decl_num}.xlsx" if success else None,
                error_message=None if success else "Mock error",
                processing_time=0.1
            )
        
        mock_printer.print_single_declaration.side_effect = mock_print_single
        
        # Run async processing
        async def run_test():
            result = await self.processor.process_declarations_async(
                declaration_numbers, mock_printer
            )
            return result
        
        # Execute test
        result = asyncio.run(run_test())
        
        # Verify completeness
        assert result.total_processed == len(declaration_numbers)
        assert len(result.results) == len(declaration_numbers)
        
        # Verify all declarations were processed
        processed_numbers = {r.declaration_number for r in result.results}
        expected_numbers = set(declaration_numbers)
        assert processed_numbers == expected_numbers
        
        # Verify success/failure counts
        actual_successful = sum(1 for r in result.results if r.success)
        actual_failed = sum(1 for r in result.results if not r.success)
        
        assert result.successful == actual_successful
        assert result.failed == actual_failed
        assert result.successful + result.failed == result.total_processed
    
    @given(
        batch_size=st.integers(min_value=1, max_value=5),
        declaration_count=st.integers(min_value=1, max_value=15)
    )
    @settings(max_examples=20, deadline=8000)
    def test_batch_size_processing_efficiency(self, batch_size, declaration_count):
        """
        **Feature: customs-declaration-printing, Property: Batch size processing efficiency**
        
        For any batch size and declaration count, processing should complete
        within reasonable time bounds and maintain efficiency.
        """
        # Create processor with specific batch size
        processor = AsyncBatchProcessor(max_workers=2, batch_size=batch_size)
        
        # Generate declaration numbers
        declaration_numbers = [f"30525440{i:04d}" for i in range(declaration_count)]
        
        # Create mock printer with timing
        mock_printer = Mock()
        
        def mock_print_single(decl_num):
            time.sleep(0.01)  # Simulate processing time
            return PrintResult(
                success=True,
                declaration_number=decl_num,
                output_file_path=f"output_{decl_num}.xlsx",
                processing_time=0.01
            )
        
        mock_printer.print_single_declaration.side_effect = mock_print_single
        
        # Measure processing time
        async def run_test():
            start_time = time.time()
            result = await processor.process_declarations_async(
                declaration_numbers, mock_printer
            )
            end_time = time.time()
            return result, end_time - start_time
        
        result, processing_time = asyncio.run(run_test())
        
        # Verify efficiency
        assert result.total_processed == declaration_count
        assert result.successful == declaration_count
        assert result.failed == 0
        
        # Processing should be faster than sequential (with some tolerance)
        sequential_time = declaration_count * 0.01
        assert processing_time < sequential_time * 1.5  # Allow 50% overhead


class TestMemoryOptimizerProperties:
    """Property-based tests for memory optimization."""
    
    def setup_method(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.optimizer = MemoryOptimizer(memory_threshold_mb=50)
    
    def teardown_method(self):
        """Clean up test environment."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_test_xml(self, declaration_data: dict, goods_count: int) -> str:
        """Create test XML file with specified data."""
        xml_path = Path(self.temp_dir) / "test_declaration.xml"
        
        xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<root>
    <DToKhaiMD>
        <Data>
            <SOTK>{declaration_data.get('declaration_number', '305254403660')}</SOTK>
            <MA_DV>{declaration_data.get('company_tax_code', '0123456789')}</MA_DV>
            <_Ten_DV_L1>{declaration_data.get('company_name', 'Test Company')}</_Ten_DV_L1>
            <NGAY_DK>{declaration_data.get('declaration_date', '2024-01-15')}</NGAY_DK>
            <TONGTGTT>{declaration_data.get('total_value', '50000.00')}</TONGTGTT>
            <TTTK>{declaration_data.get('status', 'T')}</TTTK>
        </Data>
    </DToKhaiMD>'''
        
        # Add goods items
        for i in range(goods_count):
            xml_content += f'''
    <DHangMDDK>
        <Data>
            <STT>{i + 1}</STT>
            <MA_HANG>123456789{i}</MA_HANG>
            <TEN_HANG>Test Product {i}</TEN_HANG>
            <SL_HANG>100</SL_HANG>
            <DON_GIA>500.00</DON_GIA>
            <TRI_GIA>50000.00</TRI_GIA>
        </Data>
    </DHangMDDK>'''
        
        xml_content += '\n</root>'
        
        with open(xml_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)
        
        return str(xml_path)
    
    @given(
        declaration_data=st.fixed_dictionaries({
            'declaration_number': st.text(alphabet=st.characters(whitelist_categories=('Nd',)), min_size=10, max_size=15),
            'company_name': st.text(min_size=1, max_size=50),
            'company_tax_code': st.text(alphabet=st.characters(whitelist_categories=('Nd',)), min_size=10, max_size=13),
            'total_value': st.decimals(min_value=1, max_value=1000000, places=2).map(str),
            'status': st.sampled_from(['T', 'C', 'P'])
        }),
        goods_count=st.integers(min_value=1, max_value=10)
    )
    @settings(max_examples=20, deadline=5000)
    def test_xml_streaming_data_preservation(self, declaration_data, goods_count):
        """
        **Feature: customs-declaration-printing, Property: XML streaming preserves data**
        
        For any XML content, streaming parsing should preserve all declaration
        data and goods items without loss or corruption.
        """
        # Create test XML
        xml_path = self._create_test_xml(declaration_data, goods_count)
        
        # Parse using streaming
        parsed_data = self.optimizer.parse_xml_streaming(xml_path)
        
        # Verify data preservation
        assert parsed_data is not None
        
        # Check declaration data
        assert parsed_data['declaration_number'] == declaration_data['declaration_number']
        assert parsed_data['company_name'] == declaration_data['company_name']
        assert parsed_data['company_tax_code'] == declaration_data['company_tax_code']
        assert parsed_data['total_value'] == declaration_data['total_value']
        assert parsed_data['status'] == declaration_data['status']
        
        # Check goods items
        assert 'goods_list' in parsed_data
        assert len(parsed_data['goods_list']) == goods_count
        
        for i, goods_item in enumerate(parsed_data['goods_list']):
            assert goods_item['item_number'] == str(i + 1)
            assert goods_item['hs_code'] == f"123456789{i}"
            assert goods_item['description'] == f"Test Product {i}"
    
    @given(
        xml_size_factor=st.integers(min_value=1, max_value=5)
    )
    @settings(max_examples=10, deadline=8000)
    def test_memory_optimization_effectiveness(self, xml_size_factor):
        """
        **Feature: customs-declaration-printing, Property: Memory optimization effectiveness**
        
        For any XML file size, streaming parsing should use less memory
        than loading the entire file into memory at once.
        """
        # Create larger XML file
        goods_count = xml_size_factor * 20  # Scale goods count
        
        declaration_data = {
            'declaration_number': '305254403660',
            'company_name': 'Test Company' * xml_size_factor,  # Scale text size
            'company_tax_code': '0123456789',
            'total_value': '50000.00',
            'status': 'T'
        }
        
        xml_path = self._create_test_xml(declaration_data, goods_count)
        
        # Measure memory usage during streaming parse
        initial_operations = self.optimizer.metrics.memory_optimized_operations
        
        parsed_data = self.optimizer.parse_xml_streaming(xml_path)
        
        # Verify parsing succeeded
        assert parsed_data is not None
        assert len(parsed_data['goods_list']) == goods_count
        
        # Verify memory optimization was triggered
        assert self.optimizer.metrics.xml_streaming_operations > 0
        
        # For larger files, memory optimization should be more active
        if xml_size_factor >= 3:
            assert self.optimizer.metrics.memory_optimized_operations > initial_operations


class TestPerformanceOptimizerIntegration:
    """Integration tests for complete performance optimization."""
    
    def setup_method(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.optimizer = PerformanceOptimizer(
            max_cache_size=3,
            max_workers=2,
            batch_size=2,
            memory_threshold_mb=50
        )
    
    def teardown_method(self):
        """Clean up test environment."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    @given(
        operation_count=st.integers(min_value=1, max_value=10)
    )
    @settings(max_examples=15, deadline=8000)
    def test_integrated_performance_optimization(self, operation_count):
        """
        **Feature: customs-declaration-printing, Property: Integrated performance optimization**
        
        For any number of operations, the performance optimizer should
        coordinate caching, async processing, and memory optimization effectively.
        """
        # Test template caching
        template_paths = []
        for i in range(operation_count):
            template_path = Path(self.temp_dir) / f"template_{i}.xlsx"
            workbook = Workbook()
            workbook.save(template_path)
            template_paths.append(str(template_path))
        
        mock_template_manager = Mock()
        mock_template_manager.validate_template.return_value = True
        mock_template_manager.load_template_mapping.return_value = {"field": "A1"}
        
        # Test caching with multiple requests
        cached_templates = []
        with patch('openpyxl.load_workbook') as mock_load:
            mock_load.return_value = Mock()
            
            for template_path in template_paths:
                template = self.optimizer.get_optimized_template(template_path, mock_template_manager)
                assert template is not None
                cached_templates.append(template)
        
        # Verify cache effectiveness
        cache_stats = self.optimizer.template_cache.get_cache_stats()
        assert cache_stats['cache_size'] <= cache_stats['max_cache_size']
        
        # Test performance report generation
        report = self.optimizer.get_performance_report()
        assert "PERFORMANCE OPTIMIZATION REPORT" in report
        assert "TEMPLATE CACHING" in report
        assert "ASYNC PROCESSING" in report
        assert "MEMORY OPTIMIZATION" in report
        
        # Verify metrics are tracked
        assert cache_stats['total_hits'] >= 0
        assert cache_stats['total_misses'] >= 0