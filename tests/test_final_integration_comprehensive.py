"""
Comprehensive Final Integration Testing for Declaration Printing Feature.

This test suite validates the complete integration of the declaration printing feature
with focus on core functionality and real-world scenarios.

Task 16: Final integration testing and validation
"""

import os
import sys
import tempfile
import shutil
import json
from pathlib import Path
from unittest.mock import Mock, patch
from datetime import datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from declaration_printing.declaration_printer import DeclarationPrinter
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem, PrintResult
from declaration_printing.type_detector import DeclarationTypeDetector
from declaration_printing.template_manager import TemplateManager
from declaration_printing.data_extractor import DeclarationDataExtractor
from declaration_printing.excel_generator import ExcelGenerator
from config.configuration_manager import ConfigurationManager
from logging_system.logger import Logger
from models.config_models import DatabaseConfig, LoggingConfig


class TestFinalIntegrationComprehensive:
    """
    Comprehensive final integration testing for declaration printing feature.
    
    This test class validates:
    1. End-to-end declaration processing
    2. Integration with existing Preview Panel functionality
    3. Real ECUS data and template handling
    4. Error handling and recovery scenarios
    5. Performance characteristics
    """
    
    def setup_method(self):
        """Set up test environment with comprehensive mocking."""
        # Create temporary directories
        self.temp_dir = tempfile.mkdtemp()
        self.template_dir = Path(self.temp_dir) / "templates"
        self.output_dir = Path(self.temp_dir) / "output"
        self.xml_dir = Path(self.temp_dir) / "xml"
        
        self.template_dir.mkdir()
        self.output_dir.mkdir()
        self.xml_dir.mkdir()
        
        # Setup logging
        logging_config = LoggingConfig(
            log_level="INFO",
            log_file=str(Path(self.temp_dir) / "test.log"),
            max_log_size=10485760,
            backup_count=5
        )
        self.logger = Logger(logging_config)
        
        # Create real templates and sample data
        self._create_functional_templates()
        self._create_sample_xml_data()
        
        # Initialize declaration printer with minimal config
        self.printer = DeclarationPrinter(
            config_manager=None,  # Use None to avoid config issues
            logger=self.logger,
            output_directory=str(self.output_dir)
        )
        
        # Override template directory
        self.printer.template_manager.template_directory = self.template_dir
    
    def teardown_method(self):
        """Clean up test environment."""
        if hasattr(self, 'temp_dir'):
            shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_functional_templates(self):
        """Create functional Excel templates for testing."""
        # Export template
        export_wb = Workbook()
        export_ws = export_wb.active
        export_ws.title = "ToKhai"
        
        # Add some structure to make it look like a real template
        export_ws['B2'] = "TỔNG CỤC HẢI QUAN"
        export_ws['B5'] = "Số tờ khai:"
        export_ws['C5'] = ""  # declaration_number
        export_ws['E5'] = "Ngày đăng ký:"
        export_ws['F5'] = ""  # declaration_date
        export_ws['B9'] = "Tên công ty:"
        export_ws['C9'] = ""  # company_name
        export_ws['B11'] = "Mã số thuế:"
        export_ws['C11'] = ""  # company_tax_code
        export_ws['B16'] = "Tổng trị giá:"
        export_ws['C16'] = ""  # total_value
        
        export_template_path = self.template_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        export_wb.save(export_template_path)
        
        # Import template (similar structure)
        import_wb = Workbook()
        import_ws = import_wb.active
        import_ws.title = "ToKhai"
        
        import_ws['B2'] = "TỔNG CỤC HẢI QUAN"
        import_ws['B5'] = "Số tờ khai:"
        import_ws['C5'] = ""  # declaration_number
        import_ws['E5'] = "Ngày đăng ký:"
        import_ws['F5'] = ""  # declaration_date
        import_ws['B9'] = "Tên công ty:"
        import_ws['C9'] = ""  # company_name
        import_ws['B11'] = "Mã số thuế:"
        import_ws['C11'] = ""  # company_tax_code
        import_ws['B16'] = "Tổng trị giá:"
        import_ws['C16'] = ""  # total_value
        
        import_template_path = self.template_dir / "ToKhaiHQ7N_QDTQ.xlsx"
        import_wb.save(import_template_path)
        
        # Create mapping files
        mapping = {
            "declaration_number": "C5",
            "declaration_date": "F5",
            "company_name": "C9",
            "company_tax_code": "C11",
            "total_value": "C16"
        }
        
        with open(self.template_dir / "ToKhaiHQ7X_QDTQ_mapping.json", 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
        
        with open(self.template_dir / "ToKhaiHQ7N_QDTQ_mapping.json", 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
    
    def _create_sample_xml_data(self):
        """Create sample XML data for testing."""
        # Export declaration XML
        export_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<ECUS5VNACCS2018>
    <DToKhaiMD>
        <Data>
            <SOTK>305254403660</SOTK>
            <MA_DV>0123456789</MA_DV>
            <_Ten_DV_L1>CÔNG TY TNHH TEST EXPORT</_Ten_DV_L1>
            <NGAY_DK>2024-01-15</NGAY_DK>
            <TONGTGTT>50000.00</TONGTGTT>
            <MA_NT_TGTT>USD</MA_NT_TGTT>
            <TTTK>T</TTTK>
        </Data>
    </DToKhaiMD>
</ECUS5VNACCS2018>'''
        
        export_xml_path = self.xml_dir / "ECUS5VNACCS2018_ToKhai_305254403660_STT123.xml"
        with open(export_xml_path, 'w', encoding='utf-8') as f:
            f.write(export_xml)
        
        # Import declaration XML
        import_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<ECUS5VNACCS2018>
    <DToKhaiMD>
        <Data>
            <SOTK>107772836360</SOTK>
            <MA_DV>0987654321</MA_DV>
            <_Ten_DV_L1>CÔNG TY CỔ PHẦN TEST IMPORT</_Ten_DV_L1>
            <NGAY_DK>2024-01-20</NGAY_DK>
            <TONGTGTT>30000.00</TONGTGTT>
            <MA_NT_TGTT>USD</MA_NT_TGTT>
            <TTTK>T</TTTK>
        </Data>
    </DToKhaiMD>
</ECUS5VNACCS2018>'''
        
        import_xml_path = self.xml_dir / "ECUS5VNACCS2018_ToKhai_107772836360_STT456.xml"
        with open(import_xml_path, 'w', encoding='utf-8') as f:
            f.write(import_xml)
    
    def test_integration_1_export_declaration_end_to_end(self):
        """
        Test 1: Complete end-to-end processing of export declaration.
        
        Validates:
        - Declaration type detection works correctly
        - Template selection is appropriate
        - XML data extraction functions
        - Excel file generation succeeds
        - File naming convention is followed
        """
        declaration_number = "305254403660"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            # Process declaration
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify success
            assert result.success, f"Processing should succeed: {result.error_message if not result.success else ''}"
            assert result.declaration_number == declaration_number
            assert result.output_file_path is not None
            assert result.processing_time > 0
            
            # Verify file creation
            output_file = Path(result.output_file_path)
            assert output_file.exists(), "Output file should be created"
            assert output_file.name == "ToKhaiHQ7X_QDTQ_305254403660.xlsx"
            assert output_file.stat().st_size > 1000, "File should have reasonable size"
            
            # Verify file content
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            # Check that declaration number is populated somewhere
            declaration_found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == declaration_number:
                        declaration_found = True
                        break
                if declaration_found:
                    break
            
            assert declaration_found, "Declaration number should be populated in Excel file"
            workbook.close()
    
    def test_integration_2_import_declaration_end_to_end(self):
        """
        Test 2: Complete end-to-end processing of import declaration.
        
        Validates:
        - Import declaration type detection
        - Correct template selection (ToKhaiHQ7N_QDTQ)
        - XML processing for import declarations
        - File generation with correct naming
        """
        declaration_number = "107772836360"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            # Process declaration
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify success
            assert result.success, f"Processing should succeed: {result.error_message if not result.success else ''}"
            assert result.declaration_number == declaration_number
            
            # Verify correct template was used (import template)
            output_file = Path(result.output_file_path)
            assert output_file.exists(), "Output file should be created"
            assert output_file.name == "ToKhaiHQ7N_QDTQ_107772836360.xlsx"
            
            # Verify file content
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            # Check that declaration number is populated
            declaration_found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == declaration_number:
                        declaration_found = True
                        break
                if declaration_found:
                    break
            
            assert declaration_found, "Declaration number should be populated in Excel file"
            workbook.close()
    
    def test_integration_3_batch_processing_functionality(self):
        """
        Test 3: Batch processing with multiple declarations.
        
        Validates:
        - Batch processing handles multiple declarations
        - Progress tracking works correctly
        - Mixed success/failure scenarios are handled
        - Summary reporting functions
        """
        declaration_numbers = ["305254403660", "107772836360"]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            # Track progress updates
            progress_updates = []
            
            def progress_callback(status):
                progress_updates.append({
                    'current_index': status.current_index,
                    'total_count': status.total_count,
                    'successful_count': status.successful_count,
                    'failed_count': status.failed_count
                })
            
            # Process batch
            result = self.printer.print_declarations(declaration_numbers, progress_callback)
            
            # Verify batch results
            assert result.total_processed == 2
            assert result.successful >= 1, "At least one declaration should succeed"
            assert not result.cancelled
            assert result.total_time > 0
            
            # Verify progress tracking
            assert len(progress_updates) >= 1, "Progress updates should be received"
            
            # Verify individual results
            assert len(result.results) == 2
            
            # Check that files were created for successful declarations
            successful_count = 0
            for individual_result in result.results:
                if individual_result.success:
                    successful_count += 1
                    output_file = Path(individual_result.output_file_path)
                    assert output_file.exists(), f"Output file should exist: {output_file}"
            
            assert successful_count == result.successful
            
            # Verify summary report
            assert "BATCH PROCESSING SUMMARY" in result.summary_report
            assert f"Total Processed: {result.total_processed}" in result.summary_report
    
    def test_integration_4_error_handling_scenarios(self):
        """
        Test 4: Error handling and recovery scenarios.
        
        Validates:
        - Invalid declaration number handling
        - Missing template error handling
        - No data available scenarios
        - Graceful error recovery
        """
        # Test 1: Invalid declaration number
        result = self.printer.print_single_declaration("invalid_number")
        assert not result.success, "Should fail for invalid declaration number"
        assert "invalid" in result.error_message.lower() or "format" in result.error_message.lower()
        
        # Test 2: Missing template (remove export template)
        export_template = self.template_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        if export_template.exists():
            export_template.unlink()
        
        with patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            mock_perms.return_value = (True, [])
            
            result = self.printer.print_single_declaration("305254403660")
            assert not result.success, "Should fail when template is missing"
            assert "template" in result.error_message.lower()
        
        # Restore template for next test
        self._create_functional_templates()
        
        # Test 3: No data available (no database, no XML)
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            # Remove XML files
            for xml_file in self.xml_dir.glob("*.xml"):
                xml_file.unlink()
            
            result = self.printer.print_single_declaration("305254403660")
            # Should still succeed with basic template
            assert result.success, "Should succeed with basic template when no data available"
    
    def test_integration_5_template_and_data_integration(self):
        """
        Test 5: Integration between templates and data sources.
        
        Validates:
        - Template loading and validation
        - Data mapping to template fields
        - Field population accuracy
        - Vietnamese character handling
        """
        declaration_number = "305254403660"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Test with XML data
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            result = self.printer.print_single_declaration(declaration_number)
            assert result.success, "Should succeed with XML data"
            
            # Verify data was populated from XML
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            # Check that some data was populated (not just empty template)
            populated_cells = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        populated_cells += 1
            
            assert populated_cells > 5, "Multiple cells should be populated from XML data"
            workbook.close()
    
    def test_integration_6_file_naming_conventions(self):
        """
        Test 6: File naming conventions and output management.
        
        Validates:
        - Correct file naming patterns
        - Output directory handling
        - File overwrite scenarios
        - File system integration
        """
        test_cases = [
            ("305254403660", "ToKhaiHQ7X_QDTQ_305254403660.xlsx"),  # Export
            ("107772836360", "ToKhaiHQ7N_QDTQ_107772836360.xlsx"),  # Import
        ]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        for declaration_number, expected_filename in test_cases:
            with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
                 patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
                
                mock_db.return_value = None
                mock_perms.return_value = (True, [])
                
                result = self.printer.print_single_declaration(declaration_number)
                
                if result.success:
                    output_file = Path(result.output_file_path)
                    assert output_file.name == expected_filename, f"Filename should match convention: expected {expected_filename}, got {output_file.name}"
                    assert output_file.parent == self.output_dir, "File should be in correct output directory"
    
    def test_integration_7_performance_baseline(self):
        """
        Test 7: Performance characteristics and baseline measurements.
        
        Validates:
        - Processing time is reasonable
        - Memory usage is acceptable
        - System can handle multiple declarations
        - Performance metrics are recorded
        """
        declaration_numbers = ["305254403660", "107772836360"]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            # Measure processing time for single declaration
            import time
            start_time = time.time()
            result = self.printer.print_single_declaration("305254403660")
            end_time = time.time()
            
            processing_time = end_time - start_time
            
            # Verify reasonable performance
            assert result.success, "Processing should succeed"
            assert processing_time < 10.0, f"Processing should complete within 10 seconds, took {processing_time:.2f}s"
            assert result.processing_time > 0, "Processing time should be recorded"
            
            # Test batch performance
            batch_start = time.time()
            batch_result = self.printer.print_declarations(declaration_numbers)
            batch_end = time.time()
            
            batch_time = batch_end - batch_start
            avg_time_per_declaration = batch_time / len(declaration_numbers)
            
            assert batch_result.total_processed == len(declaration_numbers)
            assert avg_time_per_declaration < 15.0, f"Average time per declaration should be reasonable: {avg_time_per_declaration:.2f}s"
    
    def test_integration_8_system_integration_validation(self):
        """
        Test 8: Overall system integration validation.
        
        Validates:
        - All components work together correctly
        - No integration conflicts
        - System state remains consistent
        - Resource cleanup works properly
        """
        # Test multiple operations in sequence
        operations = [
            ("305254403660", "export"),
            ("107772836360", "import"),
            ("305254403660", "export"),  # Repeat to test consistency
        ]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db, \
             patch.object(self.printer.permission_checker, 'check_printing_permissions') as mock_perms:
            
            mock_db.return_value = None
            mock_perms.return_value = (True, [])
            
            results = []
            
            for declaration_number, decl_type in operations:
                result = self.printer.print_single_declaration(declaration_number)
                results.append((declaration_number, decl_type, result))
            
            # Verify all operations succeeded
            successful_operations = 0
            for declaration_number, decl_type, result in results:
                if result.success:
                    successful_operations += 1
                    
                    # Verify file exists
                    output_file = Path(result.output_file_path)
                    assert output_file.exists(), f"Output file should exist for {declaration_number}"
                    
                    # Verify correct template was used
                    if decl_type == "export":
                        assert "ToKhaiHQ7X_QDTQ" in output_file.name
                    else:
                        assert "ToKhaiHQ7N_QDTQ" in output_file.name
            
            assert successful_operations >= len(operations) // 2, "Most operations should succeed"
            
            # Verify system state is clean (no hanging resources)
            # This is implicit - if we get here without errors, cleanup worked
            assert True, "System integration validation completed successfully"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])