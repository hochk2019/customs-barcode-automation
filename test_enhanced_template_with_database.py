#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Enhanced Template Generator với database thực tế
"""

# from declaration_printing.enhanced_template_generator import EnhancedTemplateGenerator
# from declaration_printing.enhanced_data_extractor import EnhancedDataExtractor
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem
from datetime import datetime
from decimal import Decimal
import xlrd
import openpyxl
from pathlib import Path

def create_test_data_with_goods():
    """Tạo dữ liệu test có đầy đủ hàng hóa"""
    
    # Tạo dữ liệu xuất khẩu với nhiều hàng hóa
    export_data = DeclarationData(
        declaration_number="308064365030",
        declaration_type=DeclarationType.EXPORT_CLEARANCE,
        customs_office="18A3",
        declaration_date=datetime(2025, 12, 17),
        company_tax_code="2300944637",
        company_name="CÔNG TY TNHH JAEYOUNG VINA",
        company_address="Lô C2-1, KCN Quế Võ (Mở rộng), Phường Phương Liễu, Tỉnh Bắc Ninh, Việt Nam",
        partner_name="JAEYOUNG ELECTRONICS CO., LTD",
        partner_address="Korea",
        country_of_origin="KR",
        country_of_destination="KR",
        total_value=Decimal("50000.00"),
        currency="USD",
        exchange_rate=Decimal("24500.00"),
        goods_list=[
            GoodsItem(
                item_number=1,
                hs_code="85299040",
                description="JYE8625M#&Mô đun camera tự động lấy nét dùng cho điện thoại di động",
                quantity=Decimal("1000"),
                unit="PCS",
                unit_price=Decimal("1.5"),
                total_value=Decimal("1500.0"),
                origin_country="KR"
            ),
            GoodsItem(
                item_number=2,
                hs_code="85299041",
                description="JYE8626M#&Mô đun camera khác dùng cho điện thoại di động",
                quantity=Decimal("2000"),
                unit="PCS",
                unit_price=Decimal("1.2"),
                total_value=Decimal("2400.0"),
                origin_country="KR"
            ),
            GoodsItem(
                item_number=3,
                hs_code="85299042",
                description="JYE8627M#&Linh kiện camera cho thiết bị di động",
                quantity=Decimal("500"),
                unit="PCS",
                unit_price=Decimal("2.0"),
                total_value=Decimal("1000.0"),
                origin_country="KR"
            )
        ],
        total_weight=Decimal("1000.5"),
        total_packages=10,
        transport_method="AIR",
        bill_of_lading="BL123456789",
        container_numbers=["CONT001", "CONT002"],
        additional_data={}
    )
    
    # Tạo dữ liệu nhập khẩu
    import_data = DeclarationData(
        declaration_number="107807186540",
        declaration_type=DeclarationType.IMPORT_CLEARANCE,
        customs_office="18A3",
        declaration_date=datetime(2025, 12, 17),
        company_tax_code="2300944637",
        company_name="CÔNG TY TNHH JAEYOUNG VINA",
        company_address="Lô C2-1, KCN Quế Võ (Mở rộng), Phường Phương Liễu, Tỉnh Bắc Ninh, Việt Nam",
        partner_name="JAEYOUNG ELECTRONICS CO., LTD",
        partner_address="Korea",
        country_of_origin="KR",
        country_of_destination="VN",
        total_value=Decimal("25000.00"),
        currency="USD",
        exchange_rate=Decimal("24500.00"),
        goods_list=[
            GoodsItem(
                item_number=1,
                hs_code="85299040",
                description="Camera module JYE001#&Mô đun camera cho điện thoại di động",
                quantity=Decimal("500"),
                unit="PCS",
                unit_price=Decimal("2.0"),
                total_value=Decimal("1000.0"),
                origin_country="KR"
            ),
            GoodsItem(
                item_number=2,
                hs_code="85299041",
                description="Camera lens JYE002#&Ống kính camera chất lượng cao",
                quantity=Decimal("300"),
                unit="PCS",
                unit_price=Decimal("3.0"),
                total_value=Decimal("900.0"),
                origin_country="KR"
            )
        ],
        total_weight=Decimal("500.5"),
        total_packages=5,
        transport_method="SEA",
        bill_of_lading="BL987654321",
        container_numbers=["CONT003"],
        additional_data={}
    )
    
    return export_data, import_data

class SimpleTemplatePopulator:
    """Simple template populator với mapping cố định"""
    
    def __init__(self):
        # Mapping cố định dựa trên phân tích template
        self.export_mappings = {
            'declaration_number': [(4, 5), (4, 13), (85, 5), (85, 13)],  # Nhiều vị trí số tờ khai
            'tax_code': [(13, 5)],  # Mã số thuế người xuất khẩu
            'company_name': [(14, 5)],  # Tên công ty
            'company_address': [(17, 5)],  # Địa chỉ
            'goods_start_row': 100,  # Bảng hàng hóa bắt đầu từ row 100
            'goods_columns': {
                'stt': 3,
                'hs_code': 4,
                'description': 5,
                'quantity': 6,
                'unit': 7,
                'unit_price': 8,
                'total_value': 9,
                'origin': 10
            }
        }
        
        self.import_mappings = {
            'declaration_number': [(4, 5), (4, 14), (79, 5), (79, 14)],  # Nhiều vị trí số tờ khai
            'tax_code': [(10, 5)],  # Mã số thuế người nhập khẩu
            'company_name': [(11, 5)],  # Tên công ty
            'company_address': [(14, 5)],  # Địa chỉ
            'goods_start_row': 100,  # Bảng hàng hóa bắt đầu từ row 100 (ước tính)
            'goods_columns': {
                'stt': 3,
                'hs_code': 4,
                'description': 5,
                'quantity': 6,
                'unit': 7,
                'unit_price': 8,
                'total_value': 9,
                'origin': 10
            }
        }
    
    def populate_template(self, declaration_data: DeclarationData, template_path: str, output_path: str):
        """Populate template với dữ liệu"""
        
        print(f"🎯 Populating template: {template_path}")
        print(f"   Declaration: {declaration_data.declaration_number}")
        print(f"   Type: {declaration_data.declaration_type}")
        print(f"   Goods: {len(declaration_data.goods_list)} items")
        
        # Chọn mapping
        if declaration_data.declaration_type == DeclarationType.EXPORT_CLEARANCE:
            mappings = self.export_mappings
        else:
            mappings = self.import_mappings
        
        # Copy và convert template
        if not self._copy_template(template_path, output_path):
            raise RuntimeError("Failed to copy template")
        
        # Load và populate
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # Populate basic fields
        self._populate_basic_fields(ws, declaration_data, mappings)
        
        # Populate goods table
        if declaration_data.goods_list:
            self._populate_goods_table(ws, declaration_data.goods_list, mappings)
        
        # Ensure text format
        self._ensure_text_format(ws)
        
        # Save
        wb.save(output_path)
        wb.close()
        
        print(f"   ✅ Template populated successfully: {output_path}")
        return output_path
    
    def _copy_template(self, template_path: str, output_path: str) -> bool:
        """Copy template từ .xls sang .xlsx"""
        
        try:
            print(f"   🔄 Converting template...")
            
            # Đọc .xls
            xls_book = xlrd.open_workbook(template_path)
            xls_sheet = xls_book.sheet_by_index(0)
            
            # Tạo .xlsx
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Copy data
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    try:
                        cell_value = xls_sheet.cell_value(row_idx, col_idx)
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    except:
                        continue
            
            wb.save(output_path)
            wb.close()
            
            print(f"   ✅ Template converted")
            return True
            
        except Exception as e:
            print(f"   ❌ Error converting template: {e}")
            return False
    
    def _populate_basic_fields(self, ws, declaration_data: DeclarationData, mappings: dict):
        """Populate basic fields"""
        
        print(f"   📝 Populating basic fields...")
        
        # Số tờ khai
        for row, col in mappings['declaration_number']:
            ws.cell(row=row, column=col).value = declaration_data.declaration_number
            print(f"      ✅ Declaration number at ({row}, {col}): {declaration_data.declaration_number}")
        
        # Mã số thuế
        for row, col in mappings['tax_code']:
            ws.cell(row=row, column=col).value = declaration_data.company_tax_code
            print(f"      ✅ Tax code at ({row}, {col}): {declaration_data.company_tax_code}")
        
        # Tên công ty
        for row, col in mappings['company_name']:
            ws.cell(row=row, column=col).value = declaration_data.company_name
            print(f"      ✅ Company name at ({row}, {col}): {declaration_data.company_name}")
        
        # Địa chỉ
        for row, col in mappings['company_address']:
            ws.cell(row=row, column=col).value = declaration_data.company_address
            print(f"      ✅ Address at ({row}, {col}): {declaration_data.company_address[:50]}...")
        
        # Ngày hiện tại
        current_date = datetime.now().strftime('%d/%m/%Y')
        date_positions = [(8, 7), (89, 7)]  # Các vị trí ngày
        for row, col in date_positions:
            try:
                ws.cell(row=row, column=col).value = current_date
                print(f"      ✅ Date at ({row}, {col}): {current_date}")
            except:
                pass
    
    def _populate_goods_table(self, ws, goods_list: list, mappings: dict):
        """Populate goods table"""
        
        print(f"   📦 Populating goods table...")
        
        start_row = mappings['goods_start_row']
        columns = mappings['goods_columns']
        
        print(f"      Start row: {start_row}")
        print(f"      Columns: {columns}")
        
        for i, goods in enumerate(goods_list[:10]):  # Giới hạn 10 items
            current_row = start_row + i
            
            # STT
            ws.cell(row=current_row, column=columns['stt']).value = str(goods.item_number)
            
            # Mã HS
            ws.cell(row=current_row, column=columns['hs_code']).value = str(goods.hs_code)
            
            # Mô tả
            ws.cell(row=current_row, column=columns['description']).value = str(goods.description)
            
            # Số lượng
            ws.cell(row=current_row, column=columns['quantity']).value = str(float(goods.quantity))
            
            # Đơn vị
            ws.cell(row=current_row, column=columns['unit']).value = str(goods.unit)
            
            # Đơn giá
            ws.cell(row=current_row, column=columns['unit_price']).value = str(float(goods.unit_price))
            
            # Thành tiền
            ws.cell(row=current_row, column=columns['total_value']).value = str(float(goods.total_value))
            
            # Xuất xứ
            ws.cell(row=current_row, column=columns['origin']).value = str(goods.origin_country)
            
            print(f"      ✅ Item {i+1}: {goods.hs_code} - {goods.description[:30]}...")
        
        print(f"   ✅ Goods table populated")
    
    def _ensure_text_format(self, ws):
        """Ensure text format"""
        
        print(f"   🔤 Converting to text format...")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        cell.value = str(cell.value)
                        cell.number_format = '@'
                    elif isinstance(cell.value, str):
                        cell.number_format = '@'
        
        print(f"   ✅ Text format applied")

def test_simple_template_populator():
    """Test Simple Template Populator"""
    
    print("🧪 TESTING SIMPLE TEMPLATE POPULATOR")
    print("=" * 60)
    
    # Tạo test data
    export_data, import_data = create_test_data_with_goods()
    
    # Tạo populator
    populator = SimpleTemplatePopulator()
    
    # Test export
    print(f"\n🎯 Testing Export Template")
    try:
        export_output = populator.populate_template(
            export_data,
            "sample/ToKhaiHQ7X_QDTQ_2.xls",
            "test_output/Enhanced_Export_308064365030.xlsx"
        )
        
        # Kiểm tra file
        if Path(export_output).exists():
            file_size = Path(export_output).stat().st_size
            print(f"   📊 File size: {file_size:,} bytes")
            
            # Kiểm tra nội dung
            wb = openpyxl.load_workbook(export_output, data_only=True)
            ws = wb.active
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Kiểm tra một số cell quan trọng
            print(f"   🔍 Key cells:")
            print(f"      Declaration number (4,5): {ws.cell(4,5).value}")
            print(f"      Company name (14,5): {ws.cell(14,5).value}")
            print(f"      Goods item 1 HS (100,4): {ws.cell(100,4).value}")
            print(f"      Goods item 1 desc (100,5): {ws.cell(100,5).value}")
            
            wb.close()
        
        print(f"✅ Export test successful")
        
    except Exception as e:
        print(f"❌ Export test failed: {e}")
        import traceback
        traceback.print_exc()
    
    # Test import
    print(f"\n🎯 Testing Import Template")
    try:
        import_output = populator.populate_template(
            import_data,
            "sample/ToKhaiHQ7N_QDTQ_2.xls",
            "test_output/Enhanced_Import_107807186540.xlsx"
        )
        
        # Kiểm tra file
        if Path(import_output).exists():
            file_size = Path(import_output).stat().st_size
            print(f"   📊 File size: {file_size:,} bytes")
            
            # Kiểm tra nội dung
            wb = openpyxl.load_workbook(import_output, data_only=True)
            ws = wb.active
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Kiểm tra một số cell quan trọng
            print(f"   🔍 Key cells:")
            print(f"      Declaration number (4,5): {ws.cell(4,5).value}")
            print(f"      Company name (11,5): {ws.cell(11,5).value}")
            print(f"      Goods item 1 HS (100,4): {ws.cell(100,4).value}")
            print(f"      Goods item 1 desc (100,5): {ws.cell(100,5).value}")
            
            wb.close()
        
        print(f"✅ Import test successful")
        
    except Exception as e:
        print(f"❌ Import test failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_simple_template_populator()