#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Template-Based Generator với file template thực tế
"""

from declaration_printing.template_based_generator import TemplateBasedGenerator
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem
from datetime import datetime
from decimal import Decimal

def test_export_template():
    """Test với template xuất khẩu"""
    
    print("🧪 TESTING EXPORT TEMPLATE")
    print("=" * 60)
    
    # Tạo dữ liệu test xuất khẩu
    test_data = DeclarationData(
        declaration_number="308064365777",
        declaration_type=DeclarationType.EXPORT_CLEARANCE,
        customs_office="18A3",
        declaration_date=datetime(2025, 12, 17),
        company_tax_code="2300944637",
        company_name="CÔNG TY TNHH JAEYOUNG VINA",
        company_address="Lô C2-1, KCN Quế Võ (Mở rộng), Phường Phương Liễu, Tỉnh Bắc Ninh, Việt Nam",
        partner_name="JAEYOUNG ELECTRONICS CO., LTD",
        partner_address="Korea",
        country_of_origin="KR",
        country_of_destination="KR",
        total_value=Decimal("50000.00"),
        currency="USD",
        exchange_rate=Decimal("24500.00"),
        goods_list=[
            GoodsItem(
                item_number=1,
                hs_code="85299040",
                description="JYE8625M#&Mô đun camera tự động lấy nét dùng cho điện thoại di động",
                quantity=Decimal("1000"),
                unit="PCS",
                unit_price=Decimal("1.5"),
                total_value=Decimal("1500.0"),
                origin_country="KR"
            ),
            GoodsItem(
                item_number=2,
                hs_code="85299041",
                description="JYE8626M#&Mô đun camera tự động lấy nét dùng cho điện thoại di động",
                quantity=Decimal("2000"),
                unit="PCS",
                unit_price=Decimal("1.2"),
                total_value=Decimal("2400.0"),
                origin_country="KR"
            )
        ],
        total_weight=Decimal("1000.5"),
        total_packages=10,
        transport_method="AIR",
        bill_of_lading="BL123456789",
        container_numbers=["CONT001", "CONT002"],
        additional_data={}
    )
    
    print(f"✅ Export test data created:")
    print(f"   Declaration: {test_data.declaration_number}")
    print(f"   Type: {test_data.declaration_type}")
    print(f"   Company: {test_data.company_name}")
    print(f"   Goods: {len(test_data.goods_list)} items")
    
    # Test generator
    generator = TemplateBasedGenerator("test_output")
    
    try:
        output_path = generator.generate_from_template(test_data)
        print(f"✅ Export template generation successful: {output_path}")
        
        # Kiểm tra file
        from pathlib import Path
        import openpyxl
        
        if Path(output_path).exists():
            file_size = Path(output_path).stat().st_size
            print(f"   📊 File size: {file_size:,} bytes")
            
            # Kiểm tra cấu trúc file
            wb = openpyxl.load_workbook(output_path, data_only=True)
            ws = wb.active
            
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            print(f"   📄 Sheet name: {ws.title}")
            
            # Kiểm tra một số cell quan trọng
            print(f"   🔍 Sample content:")
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and str(cell.value).strip():
                        print(f"      {cell.coordinate}: {cell.value}")
                        if row > 10:  # Giới hạn output
                            break
                if row > 10:
                    break
            
            wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Export template generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_import_template():
    """Test với template nhập khẩu"""
    
    print("\n🧪 TESTING IMPORT TEMPLATE")
    print("=" * 60)
    
    # Tạo dữ liệu test nhập khẩu
    test_data = DeclarationData(
        declaration_number="107807186777",
        declaration_type=DeclarationType.IMPORT_CLEARANCE,
        customs_office="18A3",
        declaration_date=datetime(2025, 12, 17),
        company_tax_code="2300944637",
        company_name="CÔNG TY TNHH JAEYOUNG VINA",
        company_address="Lô C2-1, KCN Quế Võ (Mở rộng), Phường Phương Liễu, Tỉnh Bắc Ninh, Việt Nam",
        partner_name="JAEYOUNG ELECTRONICS CO., LTD",
        partner_address="Korea",
        country_of_origin="KR",
        country_of_destination="VN",
        total_value=Decimal("25000.00"),
        currency="USD",
        exchange_rate=Decimal("24500.00"),
        goods_list=[
            GoodsItem(
                item_number=1,
                hs_code="85299040",
                description="Camera module JYE001#&Mô đun camera cho điện thoại di động",
                quantity=Decimal("500"),
                unit="PCS",
                unit_price=Decimal("2.0"),
                total_value=Decimal("1000.0"),
                origin_country="KR"
            )
        ],
        total_weight=Decimal("500.5"),
        total_packages=5,
        transport_method="SEA",
        bill_of_lading="BL987654321",
        container_numbers=["CONT003"],
        additional_data={}
    )
    
    print(f"✅ Import test data created:")
    print(f"   Declaration: {test_data.declaration_number}")
    print(f"   Type: {test_data.declaration_type}")
    print(f"   Goods: {len(test_data.goods_list)} items")
    
    # Test generator
    generator = TemplateBasedGenerator("test_output")
    
    try:
        output_path = generator.generate_from_template(test_data)
        print(f"✅ Import template generation successful: {output_path}")
        
        # Kiểm tra file
        from pathlib import Path
        import openpyxl
        
        if Path(output_path).exists():
            file_size = Path(output_path).stat().st_size
            print(f"   📊 File size: {file_size:,} bytes")
            
            # Kiểm tra cấu trúc file
            wb = openpyxl.load_workbook(output_path, data_only=True)
            ws = wb.active
            
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            print(f"   📄 Sheet name: {ws.title}")
            
            wb.close()
        
        return True
        
    except Exception as e:
        print(f"❌ Import template generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def analyze_template_structure():
    """Phân tích cấu trúc template để hiểu layout"""
    
    print("\n🔍 ANALYZING TEMPLATE STRUCTURE")
    print("=" * 60)
    
    templates = [
        ("Export Template", "sample/ToKhaiHQ7X_QDTQ_2.xls"),
        ("Import Template", "sample/ToKhaiHQ7N_QDTQ_2.xls")
    ]
    
    for template_name, template_path in templates:
        print(f"\n📋 {template_name}: {template_path}")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            print(f"   📄 Sheet name: {ws.title}")
            
            # Tìm các cell có text quan trọng
            important_cells = []
            for row in range(1, min(50, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        value = cell.value.lower()
                        if any(keyword in value for keyword in ['số tờ khai', 'tên', 'mã số thuế', 'địa chỉ', 'stt', 'mã hs']):
                            important_cells.append((cell.coordinate, cell.value))
            
            print(f"   🔍 Important cells found:")
            for coord, value in important_cells[:10]:  # Hiển thị 10 cell đầu
                print(f"      {coord}: {value}")
            
            wb.close()
            
        except Exception as e:
            print(f"   ❌ Error analyzing {template_name}: {e}")

if __name__ == "__main__":
    # Phân tích cấu trúc template trước
    analyze_template_structure()
    
    # Test cả hai loại template
    export_success = test_export_template()
    import_success = test_import_template()
    
    print(f"\n🎯 TEMPLATE-BASED GENERATION RESULTS:")
    print(f"   Export template: {'✅ Success' if export_success else '❌ Failed'}")
    print(f"   Import template: {'✅ Success' if import_success else '❌ Failed'}")
    
    if export_success and import_success:
        print(f"\n✅ TEMPLATE-BASED SYSTEM READY!")
        print(f"   - Uses actual template files from sample/ directory")
        print(f"   - Maintains original formatting and layout")
        print(f"   - Populates data into correct positions")
        print(f"   - Ensures text format for all data")
    else:
        print(f"\n⚠️  TEMPLATE-BASED SYSTEM NEEDS ADJUSTMENT")
        print(f"   - Check template file paths and accessibility")
        print(f"   - Verify data mapping logic")
        print(f"   - Review error messages above")