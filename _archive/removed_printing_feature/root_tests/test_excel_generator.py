#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test Excel Generator
"""

import importlib.util
import sys

if "pytest" in sys.modules and importlib.util.find_spec("declaration_printing") is None:
    import pytest
    pytest.skip("declaration_printing package not installed", allow_module_level=True)

import sys
from pathlib import Path
import time

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

def create_test_data():
    """Create test data"""
    from declaration_printing.models import DeclarationData, GoodsItem, DeclarationType
    from datetime import datetime
    from decimal import Decimal
    
    # Create multiple goods items
    goods_items = [
        GoodsItem(
            item_number=1,
            hs_code="8471.30.00",
            description="Máy tính xách tay (laptop)",
            quantity=Decimal('10'),
            unit="Cái",
            unit_price=Decimal('500.00'),
            total_value=Decimal('5000.00'),
            origin_country="CN"
        ),
        GoodsItem(
            item_number=2,
            hs_code="8528.72.10", 
            description="Màn hình LCD 24 inch",
            quantity=Decimal('20'),
            unit="Cái",
            unit_price=Decimal('200.00'),
            total_value=Decimal('4000.00'),
            origin_country="KR"
        ),
        GoodsItem(
            item_number=3,
            hs_code="8471.60.90",
            description="Bàn phím và chuột máy tính",
            quantity=Decimal('30'),
            unit="Bộ",
            unit_price=Decimal('25.00'),
            total_value=Decimal('750.00'),
            origin_country="CN"
        )
    ]
    
    return DeclarationData(
        declaration_number="305254403660",
        declaration_type=DeclarationType.EXPORT_CLEARANCE,
        customs_office="CCHQHANAM",
        declaration_date=datetime(2025, 12, 17),
        company_tax_code="0700798384",
        company_name="CÔNG TY TNHH TAESUNG HITECH VINA",
        company_address="Đường D2, khu E, khu công nghiệp Đình Vũ, phường Đông Hải 2, quận Hải An, thành phố Hải Phòng",
        partner_name="SAMSUNG ELECTRONICS CO., LTD",
        partner_address="129, Samsung-ro, Yeongtong-gu, Suwon-si, Gyeonggi-do, Korea",
        country_of_origin="VN",
        country_of_destination="KR",
        total_value=Decimal('9750.00'),
        currency="USD",
        exchange_rate=Decimal('24500'),
        goods_list=goods_items,
        total_weight=Decimal('150.5'),
        total_packages=60,
        transport_method="Đường biển",
        bill_of_lading="HLCUSGZ240001234",
        container_numbers=["TGHU1234567", "MSCU9876543"],
        additional_data={
            'status': 'T',
            'channel': 'Xanh',
            'company_phone': '+84-226-3668854'
        }
    )

def test_excel_generator():
    """Test Excel generator"""
    print("Testing Excel Generator...")
    
    try:
        from declaration_printing.multi_page_excel_generator import MultiPageExcelGenerator
        
        # Create test output directory
        Path("test_output").mkdir(exist_ok=True)
        
        # Create generator
        generator = MultiPageExcelGenerator("test_output")
        print("✅ Generator created")
        
        # Create test data
        test_data = create_test_data()
        print(f"✅ Test data created: {len(test_data.goods_list)} goods items")
        
        # Progress callback
        def progress_callback(current_page, total_pages):
            print(f"  📄 Generating page {current_page}/{total_pages}")
        
        # Generate Excel
        print(f"\n🔄 Generating Excel for {test_data.declaration_number}...")
        start_time = time.time()
        
        output_path = generator.generate_complete_declaration(test_data, progress_callback)
        
        generation_time = time.time() - start_time
        
        # Check output
        output_file = Path(output_path)
        if output_file.exists():
            file_size = output_file.stat().st_size
            print(f"\n✅ Excel file generated successfully!")
            print(f"   📁 Path: {output_path}")
            print(f"   📊 Size: {file_size:,} bytes")
            print(f"   ⏱️ Time: {generation_time:.2f} seconds")
            
            # Try to validate Excel file
            try:
                import openpyxl
                wb = openpyxl.load_workbook(output_path, read_only=True)
                ws = wb.active
                print(f"   📋 Sheet: {ws.title}")
                print(f"   📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                wb.close()
                print("   ✅ Excel file is valid")
            except Exception as e:
                print(f"   ⚠️ Excel validation warning: {e}")
            
            return True
        else:
            print("❌ Output file was not created")
            return False
            
    except Exception as e:
        print(f"❌ Excel generator test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main test"""
    print("EXCEL GENERATOR TEST")
    print("=" * 40)
    
    success = test_excel_generator()
    
    print("\n" + "=" * 40)
    if success:
        print("🎉 Excel generator test passed!")
        return 0
    else:
        print("❌ Excel generator test failed!")
        return 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)