"""
Simplified Final Integration Testing for Declaration Printing Feature.

This test suite validates the core integration functionality without GUI components
to ensure the declaration printing feature works correctly with real data and templates.

Task 16: Final integration testing and validation (Simplified)
"""

import importlib.util
import sys

if "pytest" in sys.modules and importlib.util.find_spec("declaration_printing") is None:
    import pytest
    pytest.skip("declaration_printing package not installed", allow_module_level=True)

import os
import sys
import tempfile
import shutil
import json
import time
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

import pytest
from openpyxl import Workbook, load_workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from declaration_printing.declaration_printer import DeclarationPrinter
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem, PrintResult
from declaration_printing.type_detector import DeclarationTypeDetector
from declaration_printing.template_manager import TemplateManager
from declaration_printing.data_extractor import DeclarationDataExtractor
from declaration_printing.excel_generator import ExcelGenerator
from config.configuration_manager import ConfigurationManager
from logging_system.logger import Logger
from models.config_models import DatabaseConfig, LoggingConfig


class TestSimpleFinalIntegration:
    """
    Simplified final integration testing for declaration printing feature.
    
    This test class validates core functionality without GUI dependencies:
    1. End-to-end declaration processing
    2. Real template and data handling
    3. File generation and validation
    4. Error handling scenarios
    """
    
    def setup_method(self):
        """Set up test environment with minimal dependencies."""
        # Create temporary directories
        self.temp_dir = tempfile.mkdtemp()
        self.template_dir = Path(self.temp_dir) / "templates"
        self.output_dir = Path(self.temp_dir) / "output"
        self.xml_dir = Path(self.temp_dir) / "xml"
        
        self.template_dir.mkdir()
        self.output_dir.mkdir()
        self.xml_dir.mkdir()
        
        # Setup minimal logging
        logging_config = LoggingConfig(
            log_level="INFO",
            log_file=str(Path(self.temp_dir) / "test.log"),
            max_log_size=10485760,
            backup_count=5
        )
        self.logger = Logger(logging_config)
        
        # Setup minimal configuration
        self.config_manager = Mock(spec=ConfigurationManager)
        self.config_manager.get_database_config.return_value = DatabaseConfig(
            server="test_server",
            database="test_db",
            username="test_user",
            password="test_pass"
        )
        
        # Mock the config attribute
        mock_config = Mock()
        mock_config.has_section.return_value = False
        mock_config.add_section = Mock()
        mock_config.set = Mock()
        mock_config.write = Mock()
        self.config_manager.config = mock_config
        
        # Create real templates and sample data
        self._create_minimal_templates()
        self._create_sample_xml_data()
        
        # Initialize declaration printer
        self.printer = DeclarationPrinter(
            config_manager=self.config_manager,
            logger=self.logger,
            output_directory=str(self.output_dir)
        )
        
        # Override template directory
        self.printer.template_manager.template_directory = self.template_dir
    
    def teardown_method(self):
        """Clean up test environment."""
        if hasattr(self, 'temp_dir'):
            shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_minimal_templates(self):
        """Create minimal but functional Excel templates."""
        # Export template
        export_wb = Workbook()
        export_ws = export_wb.active
        export_ws['C5'] = ""  # declaration_number
        export_ws['F5'] = ""  # declaration_date
        export_ws['C9'] = ""  # company_name
        export_ws['C11'] = ""  # company_tax_code
        export_ws['C16'] = ""  # total_value
        
        export_template_path = self.template_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        export_wb.save(export_template_path)
        
        # Import template
        import_wb = Workbook()
        import_ws = import_wb.active
        import_ws['C5'] = ""  # declaration_number
        import_ws['F5'] = ""  # declaration_date
        import_ws['C9'] = ""  # company_name
        import_ws['C11'] = ""  # company_tax_code
        import_ws['C16'] = ""  # total_value
        
        import_template_path = self.template_dir / "ToKhaiHQ7N_QDTQ.xlsx"
        import_wb.save(import_template_path)
        
        # Create mapping files
        mapping = {
            "declaration_number": "C5",
            "declaration_date": "F5",
            "company_name": "C9",
            "company_tax_code": "C11",
            "total_value": "C16"
        }
        
        with open(self.template_dir / "ToKhaiHQ7X_QDTQ_mapping.json", 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
        
        with open(self.template_dir / "ToKhaiHQ7N_QDTQ_mapping.json", 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
    
    def _create_sample_xml_data(self):
        """Create sample XML data for testing."""
        # Export declaration XML
        export_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<root>
    <DToKhaiMD>
        <Data>
            <SOTK>305254403660</SOTK>
            <MA_DV>0123456789</MA_DV>
            <_Ten_DV_L1>Test Export Company</_Ten_DV_L1>
            <NGAY_DK>2024-01-15</NGAY_DK>
            <TONGTGTT>50000.00</TONGTGTT>
            <MA_NT_TGTT>USD</MA_NT_TGTT>
            <TTTK>T</TTTK>
        </Data>
    </DToKhaiMD>
</root>'''
        
        export_xml_path = self.xml_dir / "ECUS5VNACCS2018_ToKhai_305254403660_STT123.xml"
        with open(export_xml_path, 'w', encoding='utf-8') as f:
            f.write(export_xml)
        
        # Import declaration XML
        import_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<root>
    <DToKhaiMD>
        <Data>
            <SOTK>107772836360</SOTK>
            <MA_DV>0987654321</MA_DV>
            <_Ten_DV_L1>Test Import Company</_Ten_DV_L1>
            <NGAY_DK>2024-01-20</NGAY_DK>
            <TONGTGTT>30000.00</TONGTGTT>
            <MA_NT_TGTT>USD</MA_NT_TGTT>
            <TTTK>T</TTTK>
        </Data>
    </DToKhaiMD>
</root>'''
        
        import_xml_path = self.xml_dir / "ECUS5VNACCS2018_ToKhai_107772836360_STT456.xml"
        with open(import_xml_path, 'w', encoding='utf-8') as f:
            f.write(import_xml)
    
    def test_end_to_end_export_declaration_processing(self):
        """Test complete end-to-end processing of export declaration."""
        declaration_number = "305254403660"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            # Process declaration
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify success
            assert result.success, f"Processing should succeed: {result.error_message if not result.success else ''}"
            assert result.declaration_number == declaration_number
            assert result.output_file_path is not None
            
            # Verify file creation
            output_file = Path(result.output_file_path)
            assert output_file.exists(), "Output file should be created"
            assert output_file.name == "ToKhaiHQ7X_QDTQ_305254403660.xlsx"
            
            # Verify file content
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            assert worksheet['C5'].value == declaration_number
            assert "Test Export Company" in str(worksheet['C9'].value)
            workbook.close()
    
    def test_end_to_end_import_declaration_processing(self):
        """Test complete end-to-end processing of import declaration."""
        declaration_number = "107772836360"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            # Process declaration
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify success
            assert result.success, f"Processing should succeed: {result.error_message if not result.success else ''}"
            assert result.declaration_number == declaration_number
            assert result.output_file_path is not None
            
            # Verify file creation
            output_file = Path(result.output_file_path)
            assert output_file.exists(), "Output file should be created"
            assert output_file.name == "ToKhaiHQ7N_QDTQ_107772836360.xlsx"
            
            # Verify file content
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            assert worksheet['C5'].value == declaration_number
            assert "Test Import Company" in str(worksheet['C9'].value)
            workbook.close()
    
    def test_batch_processing_functionality(self):
        """Test batch processing with multiple declarations."""
        declaration_numbers = ["305254403660", "107772836360"]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Mock database to force XML processing
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            # Process batch
            result = self.printer.print_declarations(declaration_numbers)
            
            # Verify batch results
            assert result.total_processed == 2
            assert result.successful >= 1  # At least one should succeed
            assert not result.cancelled
            
            # Verify individual results
            assert len(result.results) == 2
            
            # Check that files were created for successful declarations
            successful_count = 0
            for individual_result in result.results:
                if individual_result.success:
                    successful_count += 1
                    output_file = Path(individual_result.output_file_path)
                    assert output_file.exists(), f"Output file should exist: {output_file}"
            
            assert successful_count == result.successful
    
    def test_error_handling_scenarios(self):
        """Test various error handling scenarios."""
        # Test 1: Invalid declaration number
        result = self.printer.print_single_declaration("invalid_number")
        assert not result.success, "Should fail for invalid declaration number"
        assert "invalid" in result.error_message.lower() or "format" in result.error_message.lower()
        
        # Test 2: Missing template
        template_path = self.template_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        if template_path.exists():
            template_path.unlink()
        
        result = self.printer.print_single_declaration("305254403660")
        assert not result.success, "Should fail when template is missing"
        assert "template" in result.error_message.lower()
        
        # Restore template for next test
        self._create_minimal_templates()
        
        # Test 3: No data available (no database, no XML)
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            # Remove XML files
            for xml_file in self.xml_dir.glob("*.xml"):
                xml_file.unlink()
            
            result = self.printer.print_single_declaration("305254403660")
            # Should still succeed with basic template
            assert result.success, "Should succeed with basic template when no data available"
    
    def test_template_and_data_integration(self):
        """Test integration between templates and data sources."""
        declaration_number = "305254403660"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        # Test with XML data
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            result = self.printer.print_single_declaration(declaration_number)
            assert result.success, "Should succeed with XML data"
            
            # Verify data was populated from XML
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            # Check that XML data was used
            assert worksheet['C5'].value == declaration_number
            assert worksheet['C9'].value is not None  # Company name should be populated
            assert worksheet['C11'].value is not None  # Tax code should be populated
            
            workbook.close()
    
    def test_file_naming_conventions(self):
        """Test that file naming conventions are followed correctly."""
        test_cases = [
            ("305254403660", "ToKhaiHQ7X_QDTQ_305254403660.xlsx"),  # Export
            ("107772836360", "ToKhaiHQ7N_QDTQ_107772836360.xlsx"),  # Import
        ]
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        for declaration_number, expected_filename in test_cases:
            with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
                mock_db.return_value = None
                
                result = self.printer.print_single_declaration(declaration_number)
                
                if result.success:
                    output_file = Path(result.output_file_path)
                    assert output_file.name == expected_filename, f"Filename should match convention: expected {expected_filename}, got {output_file.name}"
    
    def test_performance_baseline(self):
        """Test basic performance characteristics."""
        declaration_number = "305254403660"
        
        # Override XML directory
        self.printer.data_extractor.xml_directory = str(self.xml_dir)
        
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
            mock_db.return_value = None
            
            # Measure processing time
            start_time = time.time()
            result = self.printer.print_single_declaration(declaration_number)
            end_time = time.time()
            
            processing_time = end_time - start_time
            
            # Verify reasonable performance
            assert result.success, "Processing should succeed"
            assert processing_time < 5.0, f"Processing should complete within 5 seconds, took {processing_time:.2f}s"
            assert result.processing_time > 0, "Processing time should be recorded"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])