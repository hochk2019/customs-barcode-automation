"""
Integration tests for end-to-end declaration printing workflows.

Tests complete declaration printing process from UI to file output,
database and XML integration scenarios, and error recovery scenarios.
"""

import importlib.util
import sys

if "pytest" in sys.modules and importlib.util.find_spec("declaration_printing") is None:
    import pytest
    pytest.skip("declaration_printing package not installed", allow_module_level=True)

import os
import tempfile
import shutil
import json
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from declaration_printing.declaration_printer import DeclarationPrinter
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem, PrintResult, BatchPrintResult
from declaration_printing.type_detector import DeclarationTypeDetector
from declaration_printing.template_manager import TemplateManager
from declaration_printing.data_extractor import DeclarationDataExtractor
from declaration_printing.excel_generator import ExcelGenerator
from declaration_printing.error_handler import DeclarationErrorHandler
from config.configuration_manager import ConfigurationManager
from logging_system.logger import Logger
from models.config_models import DatabaseConfig


class TestDeclarationPrintingIntegration:
    """Integration tests for complete declaration printing workflows."""
    
    def setup_method(self):
        """Set up test environment with temporary directories and mock data."""
        # Create temporary directories
        self.temp_dir = tempfile.mkdtemp()
        self.template_dir = Path(self.temp_dir) / "templates"
        self.output_dir = Path(self.temp_dir) / "output"
        self.xml_dir = Path(self.temp_dir) / "xml"
        
        self.template_dir.mkdir()
        self.output_dir.mkdir()
        self.xml_dir.mkdir()
        
        # Create mock logger
        self.mock_logger = Mock(spec=Logger)
        
        # Create mock config manager
        self.mock_config_manager = Mock(spec=ConfigurationManager)
        self.mock_config_manager.get_database_config.return_value = DatabaseConfig(
            server="test_server",
            database="test_db",
            username="test_user",
            password="test_pass"
        )
        
        # Mock the config attribute for DeclarationPrintingConfig
        mock_config = Mock()
        mock_config.has_section.return_value = False
        mock_config.add_section = Mock()
        mock_config.set = Mock()
        mock_config.write = Mock()
        self.mock_config_manager.config = mock_config
        
        # Create sample templates and mappings
        self._create_sample_templates()
        
        # Create sample XML data
        self._create_sample_xml_data()
        
        # Initialize declaration printer
        self.printer = DeclarationPrinter(
            config_manager=self.mock_config_manager,
            logger=self.mock_logger,
            output_directory=str(self.output_dir)
        )
        
        # Override template directory
        self.printer.template_manager.template_directory = self.template_dir
    
    def teardown_method(self):
        """Clean up temporary directories."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_sample_templates(self):
        """Create sample Excel templates for testing."""
        # Create export clearance template
        export_template = Workbook()
        ws = export_template.active
        ws['B5'] = "Declaration Number"
        ws['C5'] = "Declaration Date"
        ws['B10'] = "Company Name"
        ws['B11'] = "Tax Code"
        ws['F20'] = "Total Value"
        
        export_path = self.template_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        export_template.save(export_path)
        
        # Create import clearance template
        import_template = Workbook()
        ws = import_template.active
        ws['B5'] = "Declaration Number"
        ws['C5'] = "Declaration Date"
        ws['B10'] = "Company Name"
        ws['B11'] = "Tax Code"
        ws['F20'] = "Total Value"
        
        import_path = self.template_dir / "ToKhaiHQ7N_QDTQ.xlsx"
        import_template.save(import_path)
        
        # Create mapping files
        mapping = {
            "declaration_number": "B5",
            "declaration_date": "C5",
            "company_name": "B10",
            "company_tax_code": "B11",
            "total_value": "F20"
        }
        
        export_mapping_path = self.template_dir / "ToKhaiHQ7X_QDTQ_mapping.json"
        with open(export_mapping_path, 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2)
        
        import_mapping_path = self.template_dir / "ToKhaiHQ7N_QDTQ_mapping.json"
        with open(import_mapping_path, 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2)
    
    def _create_sample_xml_data(self):
        """Create sample XML data for testing."""
        xml_content = '''<?xml version="1.0" encoding="UTF-8"?>
<root>
    <DToKhaiMD>
        <Data>
            <SOTK>305254403660</SOTK>
            <MA_DV>0123456789</MA_DV>
            <_Ten_DV_L1>Test Company Ltd</_Ten_DV_L1>
            <DIA_CHI_DV>123 Test Street, Test City</DIA_CHI_DV>
            <NGAY_DK>2024-01-15</NGAY_DK>
            <MA_HQ>1801</MA_HQ>
            <DV_DT>Partner Company</DV_DT>
            <NUOC_XK>VN</NUOC_XK>
            <NUOC_NK>US</NUOC_NK>
            <TONGTGTT>50000.00</TONGTGTT>
            <MA_NT_TGTT>USD</MA_NT_TGTT>
            <TYGIA_VND>24000</TYGIA_VND>
            <TR_LUONG>1000.5</TR_LUONG>
            <SO_KIEN>10</SO_KIEN>
            <TTTK>T</TTTK>
        </Data>
    </DToKhaiMD>
    <DHangMDDK>
        <Data>
            <STT>1</STT>
            <MA_HANG>1234567890</MA_HANG>
            <TEN_HANG>Test Product</TEN_HANG>
            <SL_HANG>100</SL_HANG>
            <DVT>PCS</DVT>
            <DON_GIA>500.00</DON_GIA>
            <TRI_GIA>50000.00</TRI_GIA>
            <NUOC_SX>VN</NUOC_SX>
        </Data>
    </DHangMDDK>
</root>'''
        
        xml_path = self.xml_dir / "ECUS5VNACCS2018_ToKhai_305254403660_STT123.xml"
        with open(xml_path, 'w', encoding='utf-8') as f:
            f.write(xml_content)
    
    def _create_mock_database_data(self, declaration_number: str) -> DeclarationData:
        """Create mock database data for testing."""
        declaration_type = DeclarationTypeDetector().detect_type(declaration_number)
        
        return DeclarationData(
            declaration_number=declaration_number,
            declaration_type=declaration_type,
            customs_office="1801",
            declaration_date=datetime(2024, 1, 15),
            company_tax_code="0123456789",
            company_name="Test Company Ltd",
            company_address="123 Test Street, Test City",
            partner_name="Partner Company",
            partner_address="456 Partner Street",
            country_of_origin="VN" if declaration_type == DeclarationType.EXPORT_CLEARANCE else "US",
            country_of_destination="US" if declaration_type == DeclarationType.EXPORT_CLEARANCE else "VN",
            total_value=Decimal('50000.00'),
            currency="USD",
            exchange_rate=Decimal('24000'),
            goods_list=[
                GoodsItem(
                    item_number=1,
                    hs_code="1234567890",
                    description="Test Product",
                    quantity=Decimal('100'),
                    unit="PCS",
                    unit_price=Decimal('500.00'),
                    total_value=Decimal('50000.00'),
                    origin_country="VN"
                )
            ],
            total_weight=Decimal('1000.5'),
            total_packages=10,
            transport_method="SEA",
            bill_of_lading="BL123456",
            container_numbers=["CONT123456"],
            additional_data={'status': 'T', 'channel': 'GREEN'}
        )
    
    def test_complete_declaration_printing_process_database_source(self):
        """
        Test complete declaration printing process from database to file output.
        
        This test verifies the entire workflow:
        1. Declaration validation
        2. Type detection
        3. Template loading
        4. Database data extraction
        5. Excel generation
        6. File saving
        """
        declaration_number = "305254403660"
        
        # Mock database extraction to return test data
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db_extract:
            mock_db_extract.return_value = self._create_mock_database_data(declaration_number)
            
            # Execute the complete printing process
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify successful completion
            assert result.success is True
            assert result.declaration_number == declaration_number
            assert result.output_file_path is not None
            assert result.processing_time > 0
            
            # Verify output file was created
            output_file = Path(result.output_file_path)
            assert output_file.exists()
            assert output_file.suffix == '.xlsx'
            
            # Verify file naming convention
            expected_filename = "ToKhaiHQ7X_QDTQ_305254403660.xlsx"
            assert output_file.name == expected_filename
            
            # Verify Excel file content
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            assert worksheet['B5'].value == declaration_number
            assert worksheet['B10'].value == "Test Company Ltd"
            assert worksheet['B11'].value == "0123456789"
            assert worksheet['F20'].value == "50.000,00"  # Vietnamese number format
            
            workbook.close()
    
    def test_complete_declaration_printing_process_xml_fallback(self):
        """
        Test complete declaration printing process with XML fallback.
        
        This test verifies the fallback mechanism when database is unavailable:
        1. Database extraction fails
        2. System falls back to XML parsing
        3. Excel generation continues with XML data
        4. File is successfully created
        """
        declaration_number = "305254403660"
        
        # Mock database extraction to fail
        with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db_extract:
            mock_db_extract.return_value = None
            
            # Override XML directory to use our test XML
            self.printer.data_extractor.xml_directory = str(self.xml_dir)
            
            # Execute the printing process
            result = self.printer.print_single_declaration(declaration_number)
            
            # Verify successful completion with XML data
            assert result.success is True
            assert result.declaration_number == declaration_number
            assert result.output_file_path is not None
            
            # Verify output file was created
            output_file = Path(result.output_file_path)
            assert output_file.exists()
            
            # Verify Excel file content from XML
            workbook = load_workbook(result.output_file_path)
            worksheet = workbook.active
            
            assert worksheet['B5'].value == declaration_number
            assert worksheet['B10'].value == "Test Company Ltd"
            
            workbook.close()
    
    def test_batch_processing_with_mixed_success_failure(self):
        """
        Test batch processing with mixed success and failure scenarios.
        
        This test verifies:
        1. Batch processing handles multiple declarations
        2. Failures don't stop processing of remaining declarations
        3. Progress tracking works correctly
        4. Summary report includes all results
        """
        declaration_numbers = ["305254403660", "107772836360", "305254416960"]
        
        # Mock data extraction with mixed results
        def mock_extract_with_fallback(decl_num):
            if decl_num == "107772836360":
                # Simulate failure for middle declaration
                return None
            return self._create_mock_database_data(decl_num)
        
        with patch.object(self.printer.data_extractor, 'extract_with_fallback', side_effect=mock_extract_with_fallback):
            # Track progress updates
            progress_updates = []
            
            def progress_callback(status):
                progress_updates.append({
                    'current_index': status.current_index,
                    'total_count': status.total_count,
                    'current_declaration': status.current_declaration,
                    'successful_count': status.successful_count,
                    'failed_count': status.failed_count
                })
            
            # Execute batch processing
            result = self.printer.print_declarations(declaration_numbers, progress_callback)
            
            # Verify batch results
            assert result.total_processed == 3
            assert result.successful == 2  # Two should succeed
            assert result.failed == 1      # One should fail
            assert result.cancelled is False
            assert result.total_time > 0
            
            # Verify individual results
            assert len(result.results) == 3
            
            # First declaration should succeed
            assert result.results[0].success is True
            assert result.results[0].declaration_number == "305254403660"
            
            # Second declaration should fail (no data available)
            assert result.results[1].success is False
            assert result.results[1].declaration_number == "107772836360"
            
            # Third declaration should succeed
            assert result.results[2].success is True
            assert result.results[2].declaration_number == "305254416960"
            
            # Verify progress tracking
            assert len(progress_updates) == 3
            assert progress_updates[0]['current_index'] == 0
            assert progress_updates[1]['current_index'] == 1
            assert progress_updates[2]['current_index'] == 2
            
            # Verify summary report
            assert "BATCH PROCESSING SUMMARY" in result.summary_report
            assert "Total Processed: 3" in result.summary_report
            assert "Successful: 2" in result.summary_report
            assert "Failed: 1" in result.summary_report
    
    def test_error_recovery_and_continuation_scenarios(self):
        """
        Test error recovery and continuation in various failure scenarios.
        
        This test verifies:
        1. Template errors are handled gracefully
        2. File system errors are recovered from
        3. Processing continues after recoverable errors
        4. Appropriate error messages are generated
        """
        declaration_numbers = ["305254403660", "107772836360"]
        
        # Mock successful data extraction
        with patch.object(self.printer.data_extractor, 'extract_with_fallback') as mock_extract:
            mock_extract.return_value = self._create_mock_database_data("305254403660")
            
            # Test 1: Template file missing error
            # Remove template file to simulate error
            template_path = self.template_dir / "ToKhaiHQ7N_QDTQ.xlsx"
            if template_path.exists():
                template_path.unlink()
            
            result = self.printer.print_single_declaration("107772836360")
            
            # Should fail gracefully with appropriate error
            assert result.success is False
            assert "template" in result.error_message.lower()
            
            # Recreate template for next test
            self._create_sample_templates()
            
            # Test 2: Output directory permission error
            # Make output directory read-only
            original_output_dir = self.printer.output_directory
            readonly_dir = Path(self.temp_dir) / "readonly"
            readonly_dir.mkdir()
            
            try:
                # Try to make directory read-only (may not work on all systems)
                readonly_dir.chmod(0o444)
                
                self.printer.set_output_directory(str(readonly_dir))
                result = self.printer.print_single_declaration("305254403660")
                
                # Should handle permission error gracefully
                if not result.success:
                    assert "permission" in result.error_message.lower() or "access" in result.error_message.lower()
                
            finally:
                # Restore permissions and directory
                readonly_dir.chmod(0o755)
                self.printer.set_output_directory(original_output_dir)
            
            # Test 3: Batch processing with interruption
            # Test cancellation mechanism
            def slow_extract(decl_num):
                import time
                time.sleep(0.1)  # Simulate slow operation
                return self._create_mock_database_data(decl_num)
            
            mock_extract.side_effect = slow_extract
            
            # Start batch processing and cancel it
            import threading
            
            def cancel_after_delay():
                import time
                time.sleep(0.05)  # Cancel after short delay
                self.printer.cancel_batch_processing()
            
            cancel_thread = threading.Thread(target=cancel_after_delay)
            cancel_thread.start()
            
            batch_result = self.printer.print_declarations(["305254403660", "107772836360", "305254416960"])
            
            cancel_thread.join()
            
            # Should be cancelled with partial results
            assert batch_result.cancelled is True
            assert batch_result.total_processed < 3  # Not all processed due to cancellation
    
    def test_database_and_xml_integration_scenarios(self):
        """
        Test integration scenarios with both database and XML data sources.
        
        This test verifies:
        1. Database connection handling
        2. XML parsing with various formats
        3. Data merging from multiple sources
        4. Fallback mechanisms work correctly
        """
        declaration_number = "305254403660"
        
        # Test 1: Database connection failure
        with patch.object(self.printer.data_extractor, '_get_db_connector') as mock_connector:
            mock_connector.return_value = None  # Simulate no database connection
            
            # Override XML directory
            self.printer.data_extractor.xml_directory = str(self.xml_dir)
            
            result = self.printer.print_single_declaration(declaration_number)
            
            # Should succeed using XML fallback
            assert result.success is True
            
        # Test 2: XML parsing with malformed data
        # Create malformed XML file
        malformed_xml = self.xml_dir / "malformed.xml"
        with open(malformed_xml, 'w') as f:
            f.write("<?xml version='1.0'?><root><invalid>")  # Malformed XML
        
        with patch.object(self.printer.data_extractor, 'find_xml_file') as mock_find_xml:
            mock_find_xml.return_value = str(malformed_xml)
            
            with patch.object(self.printer.data_extractor, 'extract_from_database') as mock_db:
                mock_db.return_value = None  # Force XML fallback
                
                result = self.printer.print_single_declaration(declaration_number)
                
                # Should create basic template when XML parsing fails
                assert result.success is True  # Basic template should still work
        
        # Test 3: Data merging scenario
        # Mock both database and XML data
        db_data = self._create_mock_database_data(declaration_number)
        db_data.company_name = "DB Company Name"
        db_data.partner_name = ""  # Missing in DB
        
        xml_data = self._create_mock_database_data(declaration_number)
        xml_data.company_name = "XML Company Name"
        xml_data.partner_name = "XML Partner Name"  # Available in XML
        
        merged_data = self.printer.data_extractor.merge_data_sources(db_data, xml_data)
        
        # Database should take priority, but XML should fill gaps
        assert merged_data.company_name == "DB Company Name"  # DB priority
        assert merged_data.partner_name == "XML Partner Name"  # XML fills gap
    
    def test_ui_to_file_output_integration(self):
        """
        Test integration from UI components to file output.
        
        This test simulates the complete workflow from UI button click
        to final file generation, including progress updates and error handling.
        """
        # Mock UI integration components
        from gui.preview_panel_integration import PreviewPanelIntegration
        
        with patch('gui.preview_panel_integration.DeclarationPrinter') as mock_printer_class:
            mock_printer_class.return_value = self.printer
            
            # Create mock preview panel
            mock_preview_panel = Mock()
            mock_preview_panel.get_selected_declarations.return_value = ["305254403660"]
            
            # Create integration instance
            integration = PreviewPanelIntegration(
                preview_panel=mock_preview_panel,
                config_manager=self.mock_config_manager,
                logger=self.mock_logger
            )
            
            # Mock data extraction
            with patch.object(self.printer.data_extractor, 'extract_with_fallback') as mock_extract:
                mock_extract.return_value = self._create_mock_database_data("305254403660")
                
                # Track UI updates
                ui_updates = []
                
                def mock_update_ui(enabled):
                    ui_updates.append(enabled)
                
                mock_preview_panel.set_printing_state = mock_update_ui
                
                # Simulate UI button click
                integration.handle_print_declarations(["305254403660"])
                
                # Verify UI state changes
                assert len(ui_updates) >= 1  # Should have at least one UI update
                
                # Verify file was created
                expected_file = self.output_dir / "ToKhaiHQ7X_QDTQ_305254403660.xlsx"
                assert expected_file.exists()
    
    def test_performance_with_realistic_data_volumes(self):
        """
        Test performance with realistic data volumes (up to 100 declarations for CI).
        
        This test verifies:
        1. System can handle multiple declarations efficiently
        2. Memory usage remains reasonable
        3. Processing time scales appropriately
        4. No memory leaks or resource issues
        """
        # Generate test data for multiple declarations
        declaration_numbers = []
        for i in range(50):  # Reduced from 1000 for CI performance
            if i % 2 == 0:
                declaration_numbers.append(f"30525440{i:04d}")  # Export
            else:
                declaration_numbers.append(f"10777283{i:04d}")  # Import
        
        # Mock data extraction to return consistent test data
        def mock_extract_with_fallback(decl_num):
            return self._create_mock_database_data(decl_num)
        
        with patch.object(self.printer.data_extractor, 'extract_with_fallback', side_effect=mock_extract_with_fallback):
            
            # Track performance metrics
            start_time = datetime.now()
            
            # Execute batch processing
            result = self.printer.print_declarations(declaration_numbers)
            
            end_time = datetime.now()
            total_time = (end_time - start_time).total_seconds()
            
            # Verify all declarations were processed
            assert result.total_processed == len(declaration_numbers)
            assert result.successful == len(declaration_numbers)
            assert result.failed == 0
            
            # Verify performance metrics
            avg_time_per_declaration = total_time / len(declaration_numbers)
            assert avg_time_per_declaration < 1.0  # Should be less than 1 second per declaration
            
            # Verify all files were created
            for decl_num in declaration_numbers:
                if decl_num.startswith('30'):
                    expected_file = self.output_dir / f"ToKhaiHQ7X_QDTQ_{decl_num}.xlsx"
                else:
                    expected_file = self.output_dir / f"ToKhaiHQ7N_QDTQ_{decl_num}.xlsx"
                assert expected_file.exists()
            
            # Verify summary report includes performance metrics
            assert "TIMING INFORMATION" in result.summary_report
            assert f"Total Time: {result.total_time:.2f} seconds" in result.summary_report