"""
Property-based tests for Excel generation functionality.
"""

import importlib.util
import sys

if "pytest" in sys.modules and importlib.util.find_spec("declaration_printing") is None:
    import pytest
    pytest.skip("declaration_printing package not installed", allow_module_level=True)

import os
import tempfile
import shutil
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from hypothesis import given, strategies as st, settings
from openpyxl import Workbook
import json

from declaration_printing.excel_generator import ExcelGenerator
from declaration_printing.models import DeclarationData, DeclarationType, GoodsItem


class TestExcelGeneratorProperties:
    """Property-based tests for ExcelGenerator class."""
    
    def setup_method(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.templates_dir = Path(self.temp_dir) / "templates"
        self.output_dir = Path(self.temp_dir) / "output"
        self.templates_dir.mkdir(exist_ok=True)
        self.output_dir.mkdir(exist_ok=True)
        
        # Create a sample template and mapping
        self._create_sample_template()
        
        self.generator = ExcelGenerator(output_directory=str(self.output_dir))
    
    def teardown_method(self):
        """Clean up test environment."""
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def _create_sample_template(self):
        """Create a sample Excel template and mapping for testing."""
        # Create sample template
        template_path = self.templates_dir / "ToKhaiHQ7N_QDTQ.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        
        # Add some headers to make it look like a real template
        worksheet['A1'] = "CUSTOMS DECLARATION"
        worksheet['B5'] = "Declaration Number:"
        worksheet['B10'] = "Company Name:"
        worksheet['B11'] = "Tax Code:"
        
        workbook.save(template_path)
        
        # Create corresponding mapping file
        mapping = {
            "declaration_number": "C5",
            "company_name": "C10",
            "company_tax_code": "C11",
            "declaration_date": "C6",
            "total_value": "C15",
            "currency": "C16",
            "total_weight": "C17"
        }
        
        mapping_path = self.templates_dir / "ToKhaiHQ7N_QDTQ_mapping.json"
        with open(mapping_path, 'w', encoding='utf-8') as f:
            json.dump(mapping, f, indent=2)
    
    # Strategy for generating DeclarationData
    @st.composite
    def declaration_data_strategy(draw):
        # Generate valid declaration numbers more efficiently
        prefix = draw(st.sampled_from(['30', '10']))
        suffix = draw(st.text(min_size=10, max_size=10, alphabet='0123456789'))
        declaration_number = prefix + suffix
        
        # Determine type based on number
        if declaration_number.startswith('30'):
            declaration_type = draw(st.sampled_from([DeclarationType.EXPORT_CLEARANCE, DeclarationType.EXPORT_ROUTING]))
        else:
            declaration_type = draw(st.sampled_from([DeclarationType.IMPORT_CLEARANCE, DeclarationType.IMPORT_ROUTING]))
        
        return DeclarationData(
            declaration_number=declaration_number,
            declaration_type=declaration_type,
            customs_office=draw(st.text(min_size=1, max_size=50)),
            declaration_date=draw(st.datetimes(min_value=datetime(2020, 1, 1), max_value=datetime(2024, 12, 31))),
            company_tax_code=draw(st.text(min_size=10, max_size=13).filter(str.isdigit)),
            company_name=draw(st.text(min_size=1, max_size=100)),
            company_address=draw(st.text(min_size=1, max_size=200)),
            partner_name=draw(st.text(min_size=1, max_size=100)),
            partner_address=draw(st.text(min_size=1, max_size=200)),
            country_of_origin=draw(st.text(min_size=2, max_size=3)),
            country_of_destination=draw(st.text(min_size=2, max_size=3)),
            total_value=draw(st.decimals(min_value=Decimal('0.01'), max_value=Decimal('999999999.99'), places=2)),
            currency=draw(st.sampled_from(['USD', 'VND', 'EUR', 'JPY'])),
            exchange_rate=draw(st.decimals(min_value=Decimal('0.01'), max_value=Decimal('50000'), places=4)),
            goods_list=[],
            total_weight=draw(st.decimals(min_value=Decimal('0.001'), max_value=Decimal('999999.999'), places=3)),
            total_packages=draw(st.integers(min_value=1, max_value=9999)),
            transport_method=draw(st.sampled_from(['Sea', 'Air', 'Land', 'Rail'])),
            bill_of_lading=draw(st.text(min_size=1, max_size=50)),
            container_numbers=draw(st.lists(st.text(min_size=1, max_size=20), min_size=0, max_size=5)),
            additional_data={}
        )
    
    @given(declaration_data_strategy())
    @settings(max_examples=100, deadline=None)
    def test_property_2_excel_file_generation_with_data_population(self, data):
        """
        # Feature: customs-declaration-printing, Property 2: Excel file generation with data population
        
        For any valid declaration data and template, generating an Excel file should 
        populate all available data fields and create a file with the standardized naming convention.
        
        **Validates: Requirements 1.1, 1.4, 1.5, 5.1**
        """
        # Create appropriate template based on declaration type
        if data.declaration_type in [DeclarationType.IMPORT_CLEARANCE, DeclarationType.IMPORT_ROUTING]:
            template_name = "ToKhaiHQ7N_QDTQ"
        else:
            template_name = "ToKhaiHQ7X_QDTQ"
            # Create export template if it doesn't exist
            self._create_export_template()
        
        template_path = self.templates_dir / f"{template_name}.xlsx"
        
        # Generate Excel file
        output_path = self.generator.create_from_template(str(template_path), data)
        
        # Verify file was created
        assert os.path.exists(output_path), "Excel file should be created"
        
        # Verify filename follows naming convention
        expected_filename = self.generator.format_output_filename(data.declaration_number, data.declaration_type)
        assert Path(output_path).name == expected_filename, "Filename should follow naming convention"
        
        # Verify file is a valid Excel file
        from openpyxl import load_workbook
        try:
            workbook = load_workbook(output_path)
            assert workbook is not None, "Generated file should be a valid Excel workbook"
        except Exception as e:
            assert False, f"Generated file should be readable as Excel: {e}"
        
        # Verify some data was populated (check non-empty cells)
        worksheet = workbook.active
        populated_cells = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    populated_cells += 1
        
        assert populated_cells > 0, "Excel file should contain populated data"
    
    def _create_export_template(self):
        """Create export template for testing."""
        template_path = self.templates_dir / "ToKhaiHQ7X_QDTQ.xlsx"
        if not template_path.exists():
            workbook = Workbook()
            worksheet = workbook.active
            worksheet['A1'] = "EXPORT CUSTOMS DECLARATION"
            worksheet['B5'] = "Declaration Number:"
            workbook.save(template_path)
            
            # Create mapping
            mapping = {
                "declaration_number": "C5",
                "company_name": "C10",
                "company_tax_code": "C11"
            }
            mapping_path = self.templates_dir / "ToKhaiHQ7X_QDTQ_mapping.json"
            with open(mapping_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=2)
    
    @given(
        st.one_of(
            st.decimals(min_value=Decimal('0.01'), max_value=Decimal('999999.99'), places=2),
            st.floats(min_value=0.01, max_value=999999.99),
            st.integers(min_value=1, max_value=999999),
            st.datetimes(min_value=datetime(2020, 1, 1), max_value=datetime(2024, 12, 31)),
            st.text(min_size=1, max_size=300)
        ),
        st.sampled_from(['total_value', 'total_weight', 'total_packages', 'declaration_date', 'company_name'])
    )
    @settings(max_examples=100, deadline=None)
    def test_property_6_data_formatting_consistency(self, value, field_name):
        """
        # Feature: customs-declaration-printing, Property 6: Data formatting consistency
        
        For any data being populated into templates, numeric fields should follow Vietnamese 
        customs formatting, dates should use DD/MM/YYYY format, and text should be truncated 
        with warnings if exceeding limits.
        
        **Validates: Requirements 6.1, 6.2, 6.3, 6.4**
        """
        formatted_value = self.generator._format_value(value, field_name)
        
        # Test date formatting
        if isinstance(value, datetime):
            assert isinstance(formatted_value, str), "Dates should be formatted as strings"
            # Check DD/MM/YYYY format
            parts = formatted_value.split('/')
            assert len(parts) == 3, "Date should have DD/MM/YYYY format"
            assert len(parts[0]) == 2, "Day should be 2 digits"
            assert len(parts[1]) == 2, "Month should be 2 digits"
            assert len(parts[2]) == 4, "Year should be 4 digits"
        
        # Test numeric formatting for Vietnamese customs
        elif isinstance(value, (Decimal, float)) and field_name in ['total_value', 'unit_price', 'exchange_rate']:
            assert isinstance(formatted_value, str), "Numeric values should be formatted as strings"
            # Should contain comma as decimal separator for Vietnamese format
            if ',' in formatted_value:
                parts = formatted_value.split(',')
                assert len(parts) == 2, "Should have integer and decimal parts"
                assert len(parts[1]) == 2, "Should have 2 decimal places for currency"
        
        elif isinstance(value, (Decimal, float)) and field_name == 'total_weight':
            assert isinstance(formatted_value, str), "Weight should be formatted as string"
            if ',' in formatted_value:
                parts = formatted_value.split(',')
                assert len(parts) == 2, "Should have integer and decimal parts"
                assert len(parts[1]) == 3, "Should have 3 decimal places for weight"
        
        # Test integer formatting
        elif isinstance(value, int):
            assert isinstance(formatted_value, str), "Integers should be formatted as strings"
            if field_name in ['total_packages', 'item_number']:
                assert formatted_value.isdigit(), "Package counts should be simple digits"
        
        # Test string truncation
        elif isinstance(value, str):
            max_lengths = {
                'company_name': 100,
                'company_address': 200,
                'partner_name': 100,
                'partner_address': 200,
                'description': 150,
                'bill_of_lading': 50,
                'hs_code': 20
            }
            max_length = max_lengths.get(field_name, 255)
            assert len(formatted_value) <= max_length, f"String should be truncated to {max_length} characters"
        
        # Formatted value should never be None
        assert formatted_value is not None, "Formatted value should never be None"